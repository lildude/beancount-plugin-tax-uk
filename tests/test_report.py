"""
Tests for the tax report generation functionality.
"""

import pytest
import pandas as pd
from pathlib import Path
from beancount import loader
from beancount_plugin_tax_uk.spreadsheet_writer import write_tax_report_spreadsheet
from beancount_plugin_tax_uk.calculate_tax import generate_tax_related_events
from beancount_plugin_tax_uk.tax_report import generate_tax_report
from beancount_plugin_tax_uk.rate_converter import BeancountRateConverter
import openpyxl
from openpyxl.utils import get_column_letter


@pytest.mark.parametrize(
    "test_files",
    [
        pytest.param(
            (
                "trivial_sample.beancount",
                "test_report_trivial_sample.xlsx",
                "test_report_trivial_sample.pkl",
            ),
            id="trivial_sample",
        ),
        pytest.param(
            (
                "sample_KapJI_cgc.beancount",
                "test_report_KapJI_cgc.xlsx",
                "test_report_KapJI_cgc.pkl",
            ),
            id="sample_KapJI_cgc",
        ),
        pytest.param(
            (
                "sample_HS284_Example_3_2021.beancount",
                "test_report_HS284_Example_3_2021.xlsx",
                "test_report_HS284_Example_3_2021.pkl",
            ),
            id="sample_HS284_Example_3_2021",
        ),
        pytest.param(
            (
                "sample_HMRC_bed_and_breakfast.beancount",
                "test_report_HMRC_bed_and_breakfast.xlsx",
                "test_report_HMRC_bed_and_breakfast.pkl",
            ),
            id="sample_HMRC_bed_and_breakfast",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/WithAssetEvents.beancount",
                "report_WithAssetEvents.xlsx",
                "report_WithAssetEvents.pkl",
            ),
            id="WithAssetEvents",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/Blank.beancount",
                "report_Blank.xlsx",
                "report_Blank.pkl",
            ),
            id="Blank",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/SameDayMerge.beancount",
                "report_SameDayMerge.xlsx",
                "report_SameDayMerge.pkl",
            ),
            id="SameDayMerge",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/SameDayMergeInterleaved.beancount",
                "report_SameDayMergeInterleaved.xlsx",
                "report_SameDayMergeInterleaved.pkl",
            ),
            id="SameDayMergeInterleaved",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/WithAssetEventsMultipleYears.beancount",
                "report_WithAssetEventsMultipleYears.xlsx",
                "report_WithAssetEventsMultipleYears.pkl",
            ),
            id="WithAssetEventsMultipleYears",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/HMRCExample1.beancount",
                "report_HMRCExample1.xlsx",
                "report_HMRCExample1.pkl",
            ),
            id="HMRCExample1",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/WithAssetEventsSameDay.beancount",
                "report_WithAssetEventsSameDay.xlsx",
                "report_WithAssetEventsSameDay.pkl",
            ),
            id="WithAssetEventsSameDay",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/MultipleMatches.beancount",
                "report_MultipleMatches.xlsx",
                "report_MultipleMatches.pkl",
            ),
            id="MultipleMatches",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/BuySellAllBuyAgainCapitalReturn.beancount",
                "report_BuySellAllBuyAgainCapitalReturn.xlsx",
                "report_BuySellAllBuyAgainCapitalReturn.pkl",
            ),
            id="BuySellAllBuyAgainCapitalReturn",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/Simple.beancount",
                "report_Simple.xlsx",
                "report_Simple.pkl",
            ),
            id="Simple",
        ),
        pytest.param(
            (
                "cgtcalc_inputs_beancount/CarryLoss.beancount",
                "report_CarryLoss.xlsx",
                "report_CarryLoss.pkl",
            ),
            id="CarryLoss",
        ),
    ],
)
def test_tax_report(capture_output, tmp_path, test_files):
    """Run a tax report test with configurable parameters.

    Args:
        capture_output: Whether to capture output as reference files
        tmp_path: Temporary directory for test output
        test_files: Tuple of (ledger_file, spreadsheet_file, pickle_file)
    """
    ledger_file, spreadsheet_file, pickle_file = test_files

    # Define paths
    data_dir = Path("tests") / "data"
    ledger_path = data_dir / ledger_file
    output_dir = tmp_path
    spreadsheet_path = output_dir / spreadsheet_file
    pickle_path = output_dir / pickle_file
    ref_dir = data_dir / "output"

    # Load ledger file
    entries, errors, options = loader.load_file(str(ledger_path))
    if errors:
        pytest.fail(f"Failed to load test ledger: {errors}")

    # Generate tax related events
    tax_related_events = generate_tax_related_events(entries, options)

    # Generate tax report data
    rate_converter = BeancountRateConverter(entries)
    rows, tax_res, asset_type_mapping = generate_tax_report(
        entries,
        tax_related_events,
        rate_converter=rate_converter,
    )

    # Generate spreadsheet report
    write_tax_report_spreadsheet(
        str(spreadsheet_path), rows, tax_res, asset_type_mapping
    )

    # Store tax results DataFrame
    tax_res.to_pickle(pickle_path)

    if capture_output:
        # Store as reference files
        ref_dir.mkdir(parents=True, exist_ok=True)
        spreadsheet_path.rename(ref_dir / spreadsheet_file)
        pickle_path.rename(ref_dir / pickle_file)
    else:
        # Compare with reference files
        ref_dir = data_dir / "output"
        ref_spreadsheet = ref_dir / spreadsheet_file
        ref_pickle = ref_dir / pickle_file

        # Compare DataFrame
        ref_tax_res = pd.read_pickle(ref_pickle)
        pd.testing.assert_frame_equal(tax_res, ref_tax_res)

        # Compare spreadsheet files
        wb1 = openpyxl.load_workbook(spreadsheet_path)
        wb2 = openpyxl.load_workbook(ref_spreadsheet)
        ws1 = wb1.active
        ws2 = wb2.active

        # Compare cell values
        for row in range(1, ws1.max_row + 1):
            for col in range(1, ws1.max_column + 1):
                cell1 = ws1.cell(row=row, column=col)
                cell2 = ws2.cell(row=row, column=col)
                assert cell1.value == cell2.value, (
                    f"Spreadsheet mismatch at {get_column_letter(col)}{row}: {cell1.value} != {cell2.value}"
                )


def test_commission_split_in_s104_and_bb():
    """
    Test that commission is distributed proportionally when a SELL transaction
    is split between S104 and B&B rules, not duplicated.
    """
    from decimal import Decimal
    import tempfile
    import os

    # Create test beancount content
    beancount_content = """
option "title" "Commission Split Test"
option "operating_currency" "GBP"

1970-01-01 open Assets:Broker:Stocks
1970-01-01 open Assets:Broker:Cash
1970-01-01 open Expenses:Broker:Commissions
1970-01-01 open Income:Broker:PnL

1970-01-01 custom "uk-tax-config" "ignored-currencies" "GBP"
1970-01-01 custom "uk-tax-config" "commission-account" "Expenses:Broker:Commissions"

1970-01-01 commodity TESTSTOCK
1970-01-01 commodity GBP

; Buy 1000 shares first to create S104 pool
2023-01-01 * "Buy shares for S104 pool" #buy
  Assets:Broker:Stocks        1000 TESTSTOCK {10.00 GBP}
  Assets:Broker:Cash        -10000.00 GBP

; Sell 500 shares with commission - this should split between S104 and B&B
2023-06-01 * "Sell with commission - should split between S104 and B&B" #sell
  Assets:Broker:Stocks        -500 TESTSTOCK {} @ 12.00 GBP
  Assets:Broker:Cash          5980.00 GBP
  Expenses:Broker:Commissions   20.00 GBP
  Income:Broker:PnL

; Buy back 300 shares the next day - this should trigger B&B rules
2023-06-02 * "Buy back next day - triggers B&B" #buy
  Assets:Broker:Stocks         300 TESTSTOCK {11.00 GBP}
  Assets:Broker:Cash         -3300.00 GBP
"""

    # Write to temporary file
    with tempfile.NamedTemporaryFile(mode="w", suffix=".beancount", delete=False) as f:
        f.write(beancount_content)
        temp_file = f.name

    try:
        # Load and process the test data
        entries, errors, options_map = loader.load_file(temp_file)
        assert not errors, f"Beancount loading errors: {errors}"

        tax_events = generate_tax_related_events(entries, options_map)
        rows, tax_report_df, asset_type_mapping = generate_tax_report(
            entries, tax_events, rate_converter=BeancountRateConverter(entries)
        )

        # Filter to SELL events only
        sell_events = tax_report_df[tax_report_df["event_type"] == "Sell"]

        # Should have exactly 2 SELL rows (one B&B, one S104)
        assert len(sell_events) == 2, f"Expected 2 SELL rows, got {len(sell_events)}"

        # Check that one is B&B and one is S104 by looking at the details dict
        bb_rows = []
        s104_rows = []

        for idx, row in sell_events.iterrows():
            details = row["details"]
            if isinstance(details, dict) and details.get("rule") == "B&B":
                bb_rows.append(row)
            elif isinstance(details, dict) and details.get("rule") == "S104":
                s104_rows.append(row)

        assert len(bb_rows) == 1, f"Expected 1 B&B row, got {len(bb_rows)}"
        assert len(s104_rows) == 1, f"Expected 1 S104 row, got {len(s104_rows)}"

        bb_row = bb_rows[0]
        s104_row = s104_rows[0]

        # Extract the values for verification
        bb_allowable_cost = bb_row["allowable_cost"]
        s104_allowable_cost = s104_row["allowable_cost"]

        # Verify commission distribution
        # B&B: 300 shares out of 500 = 60% of commission = 12.00 GBP
        # Expected B&B allowable cost: (300 * 11.00) + 12.00 = 3312.00
        expected_bb_cost = Decimal("3312.00")

        # S104: 200 shares out of 500 = 40% of commission = 8.00 GBP
        # Expected S104 allowable cost: (200/1000 * 10000) + 8.00 = 2008.00
        expected_s104_cost = Decimal("2008.00")

        # Convert to Decimal for precise comparison
        actual_bb_cost = Decimal(str(bb_allowable_cost))
        actual_s104_cost = Decimal(str(s104_allowable_cost))

        assert actual_bb_cost == expected_bb_cost, (
            f"B&B allowable cost mismatch: expected {expected_bb_cost}, got {actual_bb_cost}"
        )
        assert actual_s104_cost == expected_s104_cost, (
            f"S104 allowable cost mismatch: expected {expected_s104_cost}, got {actual_s104_cost}"
        )

        # Verify total commission is exactly 20.00 (not duplicated to 40.00)
        total_commission_effect = (actual_bb_cost - Decimal("3300.00")) + (
            actual_s104_cost - Decimal("2000.00")
        )
        assert total_commission_effect == Decimal("20.00"), (
            f"Total commission effect should be 20.00, got {total_commission_effect}"
        )

    finally:
        # Clean up temp file
        os.unlink(temp_file)
